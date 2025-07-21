# la llibreria utilitzada serà openpyxl, per lo vist es la primera que he trobat que
# em permet modificar un excel existent i no crear-ne un de nou
import openpyxl
import math

fitxer = openpyxl.load_workbook(r'C:\Users\ftur.DOMINIOEXPO\Desktop\test.xlsx')

full_Dades = fitxer.worksheets[0]
# Crear una nova fulla per a les enganxines
# full_Enganxines = fitxer.create_sheet(title="Etiquetes")

# *******************************************************************
# Lectura de dades del fitxer Excel
# *******************************************************************

# Dades de la plantilla
ALTURA_NOM = 0
ALTURA_CHFUNCIO = 1
ALTURA_PRIMERCANAL = 2
COLUMNES_PER_DISPOSITIU = 2
DISPOSITIUS_PER_FILA = 7
ALTURA_CELA = 15
AMPLADA_CELA_PETITA = 2.71
AMPLADA_CELA_GRAN = 11.86

# Dades llegides
nom_dispositius = full_Dades['A2'].value
quantitat_dispositius = full_Dades['B2'].value
quantitat_canals = 0

nom_canals = []

for col in range(3, 19):  # des de la columna C (3) fins a la R (19)
    canal = full_Dades.cell(row=2, column=col).value
    if canal:
        nom_canals.append(canal)
        quantitat_canals += 1
    
full_Enganxines = fitxer.create_sheet(title=f"Etiquetes")
# Dades calculades - ajustar segons el nombre de canals d'aquesta iteració
MAX_COL = COLUMNES_PER_DISPOSITIU * DISPOSITIUS_PER_FILA
filesXetiqueta = ALTURA_PRIMERCANAL + quantitat_canals
MAX_ROW = math.ceil(quantitat_dispositius/DISPOSITIUS_PER_FILA) * filesXetiqueta

# *******************************************************************
# Ara que hem agafat les dades, podem començar a crear les etiquetes
# *******************************************************************

# FORMAT DE LES CEL·LES
# *******************************************************************

# Primer de tot ajustem el format de les cel·les per a que s'adaptin al contingut
for row in full_Enganxines.iter_rows(min_row=1, max_row=MAX_ROW, min_col=1, max_col=MAX_COL):
    for cell in row:
        cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        cell.font = openpyxl.styles.Font(name='Arial', size=8)
        cell.fill = openpyxl.styles.PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),
                                            right=openpyxl.styles.Side(style='thin'),
                                            top=openpyxl.styles.Side(style='thin'),
                                            bottom=openpyxl.styles.Side(style='thin'))

# Ajustem l'alçada de les files
for row_num in range(1, MAX_ROW + 1):
    full_Enganxines.row_dimensions[row_num].height = ALTURA_CELA

# NOTA: L'amplada de les columnes es configura més avall per ajustar-se al DIN A4

# CONFIGURACIÓ PER A IMPRESSIÓ EN DIN A4 (ABANS DEL CONTINGUT)
# *******************************************************************

# Configurar marges per a DIN A4 (en polzades) - Màxim aprofitament
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.pagebreak import Break

# Marges mínims per DIN A4 per aprofitar al màxim l'espai
full_Enganxines.page_margins = PageMargins(left=0.2, right=0.2, top=0.3, bottom=0.3)

# Configurar pàgina per DIN A4
full_Enganxines.page_setup.paperSize = 9  # DIN A4
full_Enganxines.page_setup.orientation = 'portrait'
full_Enganxines.page_setup.scale = 100

# Centrar contingut en la pàgina
full_Enganxines.print_options.horizontalCentered = True
full_Enganxines.print_options.verticalCentered = False

# Ajustar columnes al màxim ample disponible en DIN A4
# DIN A4: 210mm = 8.27" - marges laterals mínims = 7.87" d'ample útil
ample_util_pulgades = 8.27 - 0.4  # 7.87" (només 0.2" per cada costat)
ample_util_unitats = ample_util_pulgades * 8.5  # Conversió més precisa a unitats Excel

# Redistribuir l'amplada de les columnes per ocupar tot l'ample disponible
amplada_total_necessaria = (AMPLADA_CELA_PETITA + AMPLADA_CELA_GRAN) * DISPOSITIUS_PER_FILA
factor_escala = ample_util_unitats / amplada_total_necessaria

# Augmentar l'amplada de les columnes 1.49 cops
factor_escala = factor_escala * 1.49

print(f"Ample útil: {ample_util_pulgades:.2f}\" ({ample_util_unitats:.1f} unitats)")
print(f"Factor d'escala aplicat (amb augment 1.7x): {factor_escala:.2f}")

for col in range(1, MAX_COL + 1):
    col_letter = chr(64 + col) if col <= 26 else f"A{chr(64 + col - 26)}"  # Conversió manual A, B, C...
    
    if col % 2 == 0:
        # Columna par - amplada gran escalada
        nova_amplada = AMPLADA_CELA_GRAN * factor_escala
        full_Enganxines.column_dimensions[col_letter].width = nova_amplada
    else:
        # Columna impar - amplada petita escalada
        nova_amplada = AMPLADA_CELA_PETITA * factor_escala
        full_Enganxines.column_dimensions[col_letter].width = nova_amplada

# Calcular salts de pàgina per evitar tallar etiquetes
# Altura útil DIN A4: 297mm = 11.7" - marges = 10.2"
# 13 funciona perfecte menys per 9 canals

altura_util_punts = (12.8 - 1.5) * 72  # 734 punts aproximadament

files_per_pagina = int(altura_util_punts / (ALTURA_CELA * filesXetiqueta))
print(f"Files d'etiquetes per pàgina: {files_per_pagina}")

# Afegir salts de pàgina després de cada grup complet d'etiquetes que capigui en una pàgina
fila_grup = 0
while fila_grup < MAX_ROW:
    proxima_fila_salt = fila_grup + (files_per_pagina * filesXetiqueta)
    if proxima_fila_salt < MAX_ROW:
        full_Enganxines.row_breaks.append(Break(id=proxima_fila_salt))
        print(f"Salt de pàgina afegit a la fila: {proxima_fila_salt}")
    fila_grup = proxima_fila_salt

# CONTINGUT DE LES CEL·LES
# *******************************************************************

# Omplim les cel·les amb el contingut corresponent
for n in range(1, quantitat_dispositius + 1):
    columna_actual = ((n - 1) % DISPOSITIUS_PER_FILA) * COLUMNES_PER_DISPOSITIU + 1
    fila_actual = math.floor((n - 1) / DISPOSITIUS_PER_FILA) * filesXetiqueta + 1
    
    # Combinar les cel·les del nom del dispositiu
    full_Enganxines.merge_cells(start_row=fila_actual + ALTURA_NOM, start_column=columna_actual, 
                            end_row=fila_actual + ALTURA_NOM, end_column=columna_actual + 1)
    
    cell_nom = full_Enganxines.cell(row=fila_actual + ALTURA_NOM, column=columna_actual)
    cell_nom.value = nom_dispositius + str(n)
    cell_nom.font = openpyxl.styles.Font(name='Arial', size=8, bold=True)
    
    cell_ch = full_Enganxines.cell(row=fila_actual + ALTURA_CHFUNCIO, column=columna_actual)
    cell_ch.value = 'CH'
    cell_ch.font = openpyxl.styles.Font(name='Arial', size=8, bold=True)
    
    cell_funcio = full_Enganxines.cell(row=fila_actual + ALTURA_CHFUNCIO, column=columna_actual + 1)
    cell_funcio.value = 'FUNCIO'
    cell_funcio.font = openpyxl.styles.Font(name='Arial', size=8, bold=True)
    
    for canal in range(1, quantitat_canals + 1):
        full_Enganxines.cell(row=fila_actual + ALTURA_PRIMERCANAL + canal - 1, column=columna_actual).value = str(canal)
        full_Enganxines.cell(row=fila_actual + ALTURA_PRIMERCANAL + canal - 1, column=columna_actual + 1).value = nom_canals[canal - 1]

# *******************************************************************
# Finalment, guardem i tanquem el fitxer Excel
# *******************************************************************

# Guardar els canvis al fitxer Excel
fitxer.save(r'C:\Users\ftur.DOMINIOEXPO\Desktop\test.xlsx')
print("Fitxer guardat correctament!")

# Opcional: Tancar el fitxer
fitxer.close()